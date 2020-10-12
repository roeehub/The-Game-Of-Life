import openpyxl


def log_run(curr_run, name, cell_size):
    """
    Log a given run

    Log format:
        Each run is saved as a row in assets/cool_runs.xlsx, such that:
        new_row = [seed name, cell_size, first cell, ..., last cell, -1]

    """
    file = 'assets/cool_runs.xlsx'
    new_row = []
    for cell in curr_run:
        string = "{},{}".format(str(cell[0]), str(cell[1]))
        new_row.append(string)

    # open the sheet
    wb = openpyxl.load_workbook(filename=file)
    ws = wb['Sheet1']
    row = ws.max_row + 1

    ws.cell(row=row, column=1, value=name)
    ws.cell(row=row, column=2, value=cell_size)
    last_col = 0
    for col, entry in enumerate(new_row, start=3):
        ws.cell(row=row, column=col, value=entry)
        last_col = col
    ws.cell(row=row, column=last_col+1, value=-1)

    try:
        wb.save(file)
    except PermissionError:
        pass


def get_run_arr_by_name(name):
    if not name:
        return [-1]
    # open the sheet
    file = 'assets/cool_runs.xlsx'
    wb = openpyxl.load_workbook(filename=file)
    ws = wb['Sheet1']

    for row in ws.iter_rows(min_row=2):
        if row[0].value == name:
            run_arr_str = [row[i].value for i in range(len(row))]
            run_arr_tup = [int(run_arr_str[1])]  # cell size first
            for index, element in enumerate(run_arr_str):
                if index > 1:
                    if run_arr_str[index] == -1:
                        break
                    run_arr_tup.append(tuple(map(int, element.split(','))))
            return run_arr_tup

    return [-1]


def get_name_by_index(index):
    # open the sheet
    file = 'assets/cool_runs.xlsx'
    wb = openpyxl.load_workbook(filename=file)
    ws = wb['Sheet1']

    row = ws[index]
    return row[0].value


def check_name_exists(name):
    # open the sheet
    file = 'assets/cool_runs.xlsx'
    wb = openpyxl.load_workbook(filename=file)
    ws = wb['Sheet1']

    for row in ws.iter_rows(min_row=2):
        if row[0].value == name:
            return True
    return False

